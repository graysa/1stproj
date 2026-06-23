import datetime
from io import BytesIO
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from django.shortcuts import render, get_object_or_404
from attendance.decorators import group_login_required, staff_required
from attendance.models import CareGroup, Member, MeetingDate, AttendanceRecord, Visitor


def _default_date_range():
    today = datetime.date.today()
    return today - datetime.timedelta(days=91), today


def _parse_date_params(request):
    try:
        from_date = datetime.date.fromisoformat(request.GET.get('from', ''))
    except ValueError:
        from_date = None
    try:
        to_date = datetime.date.fromisoformat(request.GET.get('to', ''))
    except ValueError:
        to_date = None
    if not from_date or not to_date:
        from_date, to_date = _default_date_range()
    return from_date, to_date


@group_login_required
def analytics(request):
    group = get_object_or_404(CareGroup, pk=request.session['group_id'])
    return render(request, 'attendance/analytics.html', {'group': group})


@group_login_required
def analytics_data(request):
    group = get_object_or_404(CareGroup, pk=request.session['group_id'])
    from_date, to_date = _parse_date_params(request)

    meetings = MeetingDate.objects.filter(
        group=group, date__gte=from_date, date__lte=to_date
    ).order_by('date')
    meeting_ids = list(meetings.values_list('pk', flat=True))
    meeting_count = len(meeting_ids)

    weekly_labels, weekly_totals = [], []
    for meeting in meetings:
        present = AttendanceRecord.objects.filter(
            meeting_date=meeting, is_present=True
        ).count()
        visitors = Visitor.objects.filter(meeting_date=meeting).count()
        weekly_labels.append(str(meeting.date))
        weekly_totals.append(present + visitors)

    active_members = Member.objects.filter(group=group, is_active=True)
    member_rates = []
    for member in active_members:
        attended = AttendanceRecord.objects.filter(
            meeting_date_id__in=meeting_ids, member=member, is_present=True
        ).count()
        rate = round((attended / meeting_count) * 100, 1) if meeting_count > 0 else 0.0
        member_rates.append({
            'name': member.name,
            'rate': rate,
            'attended': attended,
            'total': meeting_count,
        })

    return JsonResponse({
        'weekly': {'labels': weekly_labels, 'totals': weekly_totals},
        'member_rates': member_rates,
        'from_date': str(from_date),
        'to_date': str(to_date),
    })


@staff_required
def admin_dashboard(request):
    groups = CareGroup.objects.all().order_by('name')
    return render(request, 'attendance/admin_dashboard.html', {'groups': groups})


@staff_required
def admin_dashboard_data(request):
    from_date, to_date = _parse_date_params(request)
    groups = CareGroup.objects.all().order_by('name')
    result = []

    for group in groups:
        meetings = MeetingDate.objects.filter(
            group=group, date__gte=from_date, date__lte=to_date
        ).order_by('date')
        meeting_ids = list(meetings.values_list('pk', flat=True))
        meeting_count = len(meeting_ids)

        weekly_labels, weekly_totals = [], []
        for meeting in meetings:
            present = AttendanceRecord.objects.filter(
                meeting_date=meeting, is_present=True
            ).count()
            visitors = Visitor.objects.filter(meeting_date=meeting).count()
            weekly_labels.append(str(meeting.date))
            weekly_totals.append(present + visitors)

        active_member_count = Member.objects.filter(group=group, is_active=True).count()
        total_attended = AttendanceRecord.objects.filter(
            meeting_date_id__in=meeting_ids, is_present=True
        ).count()
        denominator = meeting_count * active_member_count
        overall_rate = round((total_attended / denominator) * 100, 1) if denominator > 0 else 0.0

        result.append({
            'name': group.name,
            'weekly': {'labels': weekly_labels, 'totals': weekly_totals},
            'overall_rate': overall_rate,
            'meeting_count': meeting_count,
        })

    return JsonResponse({
        'groups': result,
        'from_date': str(from_date),
        'to_date': str(to_date),
    })


def _calc_streaks(presences):
    """presences: list of bools ordered oldest → newest."""
    current = 0
    for p in reversed(presences):
        if p:
            current += 1
        else:
            break
    longest = run = 0
    for p in presences:
        run = run + 1 if p else 0
        longest = max(longest, run)
    last6 = ''.join('●' if p else '○' for p in presences[-6:])
    return current, longest, last6


def _xl_styles():
    thin = Side(style='thin', color='DDDDDD')
    return {
        'border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'center': Alignment(horizontal='center', vertical='center'),
        'center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left': Alignment(horizontal='left', vertical='center'),
    }


def _hcell(ws, r, c, val, s):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', fgColor='007AFF')
    cell.alignment = s['center_wrap']
    cell.border = s['border']
    return cell


def _rate_style(cell, rate):
    if rate >= 75:
        cell.font = Font(bold=True, color='1A7A34', size=10)
        cell.fill = PatternFill('solid', fgColor='E8F9ED')
    elif rate >= 50:
        cell.font = Font(bold=True, color='A05C00', size=10)
        cell.fill = PatternFill('solid', fgColor='FFF3E0')
    else:
        cell.font = Font(bold=True, color='C0251C', size=10)
        cell.fill = PatternFill('solid', fgColor='FFE8E8')


@staff_required
def export_csv(request):
    wb = Workbook()
    wb.remove(wb.active)
    s = _xl_styles()
    groups = list(CareGroup.objects.all().order_by('name'))

    # ── Overview tab ─────────────────────────────────────────────────────────
    ws_ov = wb.create_sheet('Overview', 0)
    ws_ov.sheet_view.showGridLines = False
    ov_headers = ['Care Group', 'Active Members', 'Total Meetings',
                  'Avg Attendance', 'Overall Rate', 'Best Current Streak']
    col_widths = [22, 16, 16, 16, 14, 20]
    for c, (h, w) in enumerate(zip(ov_headers, col_widths), 1):
        _hcell(ws_ov, 1, c, h, s)
        ws_ov.column_dimensions[get_column_letter(c)].width = w
    ws_ov.row_dimensions[1].height = 30

    for r_idx, group in enumerate(groups, 2):
        members = list(Member.objects.filter(group=group, is_active=True).order_by('name'))
        meetings = list(MeetingDate.objects.filter(group=group).order_by('date'))
        records = {
            (r['meeting_date_id'], r['member_id']): r['is_present']
            for r in AttendanceRecord.objects.filter(
                meeting_date__group=group
            ).values('meeting_date_id', 'member_id', 'is_present')
        }
        n_m = len(meetings)
        n_mem = len(members)
        total_present = sum(
            1 for mt in meetings for mem in members
            if records.get((mt.pk, mem.pk), False)
        )
        avg_att = round(total_present / n_m, 1) if n_m else 0
        overall_rate = round((total_present / (n_m * n_mem)) * 100, 1) if n_m and n_mem else 0
        best_streak = 0
        for mem in members:
            presences = [records.get((mt.pk, mem.pk), False) for mt in meetings]
            cur, _, _ = _calc_streaks(presences)
            best_streak = max(best_streak, cur)

        row_vals = [group.name, n_mem, n_m, avg_att, overall_rate, best_streak]
        bg = 'F9F9FB' if r_idx % 2 == 0 else 'FFFFFF'
        for c, val in enumerate(row_vals, 1):
            cell = ws_ov.cell(row=r_idx, column=c, value=val)
            cell.alignment = s['center'] if c > 1 else s['left']
            cell.border = s['border']
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.font = Font(size=10, bold=(c == 1))
            if c == 5:
                cell.value = f'{val}%'
                _rate_style(cell, overall_rate)
            if c == 6 and val >= 4:
                cell.font = Font(bold=True, color='FF6B00', size=10)
    ws_ov.freeze_panes = 'A2'

    # ── Visitors tab ─────────────────────────────────────────────────────────
    ws_vis = wb.create_sheet('Visitors')
    ws_vis.sheet_view.showGridLines = False
    for c, (h, w) in enumerate(
        zip(['Date', 'Care Group', 'Name', 'Note', 'Times Visited'],
            [14, 20, 20, 32, 14]), 1
    ):
        _hcell(ws_vis, 1, c, h, s)
        ws_vis.column_dimensions[get_column_letter(c)].width = w
    ws_vis.row_dimensions[1].height = 30

    all_visitors = list(
        Visitor.objects.select_related('meeting_date__group')
        .order_by('meeting_date__group__name', 'meeting_date__date', 'name')
    )
    visit_counts = {}
    for v in all_visitors:
        key = (v.meeting_date.group_id, v.name.lower())
        visit_counts[key] = visit_counts.get(key, 0) + 1

    for r_idx, v in enumerate(all_visitors, 2):
        key = (v.meeting_date.group_id, v.name.lower())
        bg = 'F9F9FB' if r_idx % 2 == 0 else 'FFFFFF'
        for c, val in enumerate(
            [v.meeting_date.date.strftime('%d %b %Y'),
             v.meeting_date.group.name, v.name,
             v.note or '', visit_counts[key]], 1
        ):
            cell = ws_vis.cell(row=r_idx, column=c, value=val)
            cell.alignment = s['center'] if c in (1, 5) else s['left']
            cell.border = s['border']
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.font = Font(size=10)
    ws_vis.freeze_panes = 'A2'

    # ── Per-group tabs ────────────────────────────────────────────────────────
    for group in groups:
        ws = wb.create_sheet(title=group.name[:31])
        ws.sheet_view.showGridLines = False

        members = list(Member.objects.filter(group=group).order_by('name'))
        meetings = list(MeetingDate.objects.filter(group=group).order_by('date'))

        if not meetings:
            ws['A1'] = 'No meetings recorded yet.'
            continue

        records = {
            (r['meeting_date_id'], r['member_id']): r['is_present']
            for r in AttendanceRecord.objects.filter(
                meeting_date__group=group
            ).values('meeting_date_id', 'member_id', 'is_present')
        }
        visitor_counts = {}
        for v in Visitor.objects.filter(
            meeting_date__group=group
        ).values('meeting_date_id'):
            mid = v['meeting_date_id']
            visitor_counts[mid] = visitor_counts.get(mid, 0) + 1

        n_mem = len(members)
        vis_col = n_mem + 2
        total_col = n_mem + 3

        # Header row
        ws.column_dimensions['A'].width = 16
        _hcell(ws, 1, 1, 'Date', s)
        for i, m in enumerate(members):
            col = i + 2
            _hcell(ws, 1, col, m.name, s)
            ws.column_dimensions[get_column_letter(col)].width = max(len(m.name) + 2, 10)
        _hcell(ws, 1, vis_col, 'Visitors', s)
        _hcell(ws, 1, total_col, 'Total Present', s)
        ws.column_dimensions[get_column_letter(vis_col)].width = 10
        ws.column_dimensions[get_column_letter(total_col)].width = 14
        ws.row_dimensions[1].height = 36

        # Data rows with month grouping
        current_row = 2
        current_month = None
        data_rows = []
        chart_dates = []
        chart_totals = []

        for meeting in meetings:
            m_month = (meeting.date.year, meeting.date.month)
            if m_month != current_month:
                current_month = m_month
                label = meeting.date.strftime('%B %Y').upper()
                cell = ws.cell(row=current_row, column=1, value=f'  {label}')
                cell.font = Font(bold=True, color='6E6E73', size=9)
                cell.fill = PatternFill('solid', fgColor='E8E8ED')
                cell.alignment = s['left']
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=total_col
                )
                ws.row_dimensions[current_row].height = 18
                current_row += 1

            date_cell = ws.cell(row=current_row, column=1,
                                value=meeting.date.strftime('%d %b %Y'))
            date_cell.font = Font(bold=True, size=10)
            date_cell.fill = PatternFill('solid', fgColor='F2F2F7')
            date_cell.alignment = s['center']
            date_cell.border = s['border']

            member_total = 0
            for i, member in enumerate(members):
                col = i + 2
                is_present = records.get((meeting.pk, member.pk), False)
                val = 1 if is_present else 0
                member_total += val
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.alignment = s['center']
                cell.border = s['border']
                if val == 1:
                    cell.fill = PatternFill('solid', fgColor='E8F9ED')
                    cell.font = Font(color='1A7A34', bold=True, size=10)
                else:
                    cell.font = Font(color='C7C7CC', size=10)

            v_count = visitor_counts.get(meeting.pk, 0)
            vcell = ws.cell(row=current_row, column=vis_col,
                            value=v_count if v_count else '')
            vcell.alignment = s['center']
            vcell.border = s['border']

            total = member_total + v_count
            tcell = ws.cell(row=current_row, column=total_col, value=total)
            tcell.alignment = s['center']
            tcell.border = s['border']
            tcell.font = Font(bold=True, size=10)

            data_rows.append(current_row)
            chart_dates.append(meeting.date.strftime('%d %b'))
            chart_totals.append(total)
            current_row += 1

        # ── Summary section ───────────────────────────────────────────────
        current_row += 1
        n_meetings = len(data_rows)

        def label_cell(ws, row, text):
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill('solid', fgColor='E8E8ED')
            c.alignment = s['left']
            c.border = s['border']

        # Total attended
        label_cell(ws, current_row, 'Total Attended')
        for i, member in enumerate(members):
            col = i + 2
            attended = sum(1 for mt in meetings if records.get((mt.pk, member.pk), False))
            cell = ws.cell(row=current_row, column=col, value=attended)
            cell.font = Font(bold=True, size=10)
            cell.alignment = s['center']
            cell.border = s['border']
        current_row += 1

        # Attendance rate
        label_cell(ws, current_row, 'Attendance Rate')
        for i, member in enumerate(members):
            col = i + 2
            attended = sum(1 for mt in meetings if records.get((mt.pk, member.pk), False))
            rate = round((attended / n_meetings) * 100) if n_meetings else 0
            cell = ws.cell(row=current_row, column=col, value=f'{rate}%')
            cell.alignment = s['center']
            cell.border = s['border']
            _rate_style(cell, rate)
        current_row += 1

        # Current streak
        label_cell(ws, current_row, 'Current Streak 🔥')
        for i, member in enumerate(members):
            col = i + 2
            presences = [records.get((mt.pk, member.pk), False) for mt in meetings]
            cur, _, _ = _calc_streaks(presences)
            cell = ws.cell(row=current_row, column=col, value=cur)
            cell.alignment = s['center']
            cell.border = s['border']
            if cur >= 8:
                cell.font = Font(bold=True, color='FF6B00', size=11)
                cell.fill = PatternFill('solid', fgColor='FFF3E0')
            elif cur >= 4:
                cell.font = Font(bold=True, color='FF9500', size=10)
            else:
                cell.font = Font(size=10)
        current_row += 1

        # Longest streak
        label_cell(ws, current_row, 'Longest Streak 🏆')
        for i, member in enumerate(members):
            col = i + 2
            presences = [records.get((mt.pk, member.pk), False) for mt in meetings]
            _, longest, _ = _calc_streaks(presences)
            cell = ws.cell(row=current_row, column=col, value=longest)
            cell.alignment = s['center']
            cell.border = s['border']
            cell.font = Font(bold=(longest >= 4), size=10)
        current_row += 1

        # Streak history (last 6 weeks)
        label_cell(ws, current_row, 'Last 6 Weeks')
        for i, member in enumerate(members):
            col = i + 2
            presences = [records.get((mt.pk, member.pk), False) for mt in meetings]
            _, _, history = _calc_streaks(presences)
            cell = ws.cell(row=current_row, column=col, value=history)
            cell.alignment = s['center']
            cell.border = s['border']
            cell.font = Font(size=12)
        current_row += 1

        ws.freeze_panes = 'B2'

        # ── Bar chart (clean data to the side) ───────────────────────────
        if chart_totals:
            chart_start_col = total_col + 2
            chart_start_row = 1
            # Write clean chart data (no month dividers)
            ws.cell(row=chart_start_row, column=chart_start_col, value='Date').font = Font(bold=True, size=9, color='AAAAAA')
            ws.cell(row=chart_start_row, column=chart_start_col + 1, value='Present').font = Font(bold=True, size=9, color='AAAAAA')
            for i, (d, t) in enumerate(zip(chart_dates, chart_totals)):
                ws.cell(row=chart_start_row + 1 + i, column=chart_start_col, value=d)
                ws.cell(row=chart_start_row + 1 + i, column=chart_start_col + 1, value=t)

            chart = BarChart()
            chart.type = 'col'
            chart.title = f'Weekly Attendance — {group.name}'
            chart.y_axis.title = 'Present'
            chart.style = 10
            chart.height = 12
            chart.width = 22
            chart.y_axis.numFmt = '0'
            chart.shape = 4

            data_ref = Reference(ws,
                min_col=chart_start_col + 1,
                min_row=chart_start_row,
                max_row=chart_start_row + len(chart_totals))
            cats = Reference(ws,
                min_col=chart_start_col,
                min_row=chart_start_row + 1,
                max_row=chart_start_row + len(chart_totals))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)

            # Hide the raw chart data rows (light gray font)
            for i in range(len(chart_totals) + 1):
                for cc in [chart_start_col, chart_start_col + 1]:
                    cell = ws.cell(row=chart_start_row + i, column=cc)
                    cell.font = Font(color='EEEEEE', size=8)

            anchor_col = get_column_letter(chart_start_col)
            anchor_row = len(chart_totals) + 4
            ws.add_chart(chart, f'{anchor_col}{anchor_row}')

    # ── Write response ────────────────────────────────────────────────────────
    filename = f'attendance_{datetime.date.today().isoformat()}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    buffer = BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response
